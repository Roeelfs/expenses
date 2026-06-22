"""
Parse two people's expense statements and generate an HTML report.

Every personal value — names, amounts, file locations, personal-name tokens —
comes from config.py (copy config.example.py to start). This file is data-free.
"""
from __future__ import annotations
import csv, json, os, re, glob, subprocess, sys, html
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent          # repo root (parent of engine/)
sys.path.insert(0, str(ROOT))                          # so `import config` works
sys.path.insert(0, str(Path(__file__).resolve().parent))  # so `import render` works
try:
    import config as C
except ModuleNotFoundError:
    import importlib.util as _ilu, pathlib as _pl, sys as _sys
    _spec = _ilu.spec_from_file_location("config", _pl.Path(__file__).resolve().parents[1] / "config.example.py")
    C = _ilu.module_from_spec(_spec)
    _sys.modules["config"] = C
    _spec.loader.exec_module(C)
import render
import store as _store, classify_context as _cc

assert [p.id for p in C.PEOPLE] == ['person_a', 'person_b'], \
    "config.PEOPLE must define two people with ids 'person_a' and 'person_b' (change `label`, not `id`)."

OUTDIR = ROOT / C.OUTPUT_DIR
OUTDIR.mkdir(exist_ok=True)

PERIOD_START         = C.PERIOD_START
PERIOD_END           = C.PERIOD_END
MONTHS_COUNT         = C.MONTHS_COUNT
RENT_SPLIT           = C.SPLIT
RENT_PER_MONTH_TOTAL = C.RENT_PER_MONTH_TOTAL
RENT_AMOUNT          = C.RENT_SETTLEMENT_AMOUNT     # one person's rent share, transferred to settle
RENTAL_INCOME_AMOUNT = C.RENTAL_INCOME_AMOUNT
SALARY_MIN_AMOUNT    = C.SALARY_MIN_AMOUNT
CURRENCY             = C.CURRENCY

# Generic bank/transfer descriptors (banking structure, not personal names).
SALARY_MERCHANT_RE = re.compile(r'העברת|העברה ?דיגיט|בנק[א-ת ]*אוצר|בנק[א-ת ]*דיסקונט|בנק[א-ת ]*הפועלים|בנק[א-ת ]*לאומי')


def _has_token(text: str, tokens) -> bool:
    """True if any config name-token appears in `text` (space-insensitive match)."""
    if not text:
        return False
    squashed = text.replace(' ', '')
    return any(tok and (tok in text or tok.replace(' ', '') in squashed) for tok in tokens)

# ─── Hebrew normalization ────────────────────────────────────────────────────

def strip_rtl(s: str) -> str:
    if not s: return s
    # Remove Unicode bidi markers + nbsp
    return re.sub(r'[‎‏‪-‮⁦-⁩ ]', '', s).strip()

def parse_amount(v) -> float:
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).replace(',', '').replace('₪', '').replace('$', '').strip()
    try: return float(s)
    except: return 0.0

def parse_date_he(s: str) -> date | None:
    """Parses DD-MM-YYYY or DD/MM/YY."""
    if not s: return None
    s = s.strip()
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%d/%m/%y', '%d-%m-%y'):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None

# ─── Moneytor source adapter (§4.2) ──────────────────────────────────────────
_BANK_TYPES = {"CHECKING", "SAVINGS", "BANK"}
_BANK_CATEGORIES = {"BANK_TRANSFER"}

def _account_kind(raw: dict) -> str:
    if raw.get("type", "").upper() in _BANK_TYPES or raw.get("category", "") in _BANK_CATEGORIES:
        return "bank"
    return "card"

def _derive_bank_kind(merchant: str, signed: float, raw: dict) -> str:
    """Consolidated bank_kind (was duplicated across the deleted loaders). Tolerance on
    ALL amount comparisons. Re-tune cheque tokens vs real Moneytor strings at the gate."""
    desc = f"{merchant} {raw.get('extra_info') or ''}"
    is_debit = signed < 0
    amt = abs(signed)
    def near(a, b): return bool(b) and abs(a - b) < 0.01
    if is_debit and _has_token(desc, C.LANDLORD_TOKENS) and (near(amt, C.RENT_PER_MONTH_TOTAL) or "שיק" in desc):
        return "landlord_rent"
    if "שיק" in desc and "עבר זמ" in desc:
        return "non_income_credit"
    if (not is_debit) and _has_token(desc, C.SALARY_KEYWORDS) and amt >= C.SALARY_MIN_AMOUNT:
        return "salary"
    if (not is_debit) and near(amt, C.RENTAL_INCOME_AMOUNT):
        return "rental_income"
    if _has_token(desc, C.COUPLE_NAME_TOKENS):
        return "couple_transfer"
    if _has_token(desc, getattr(C, "SELF_TRANSFER_TOKENS", [])):
        return "non_income_credit"
    return "other"

def map_moneytor(raw: dict, owner: str) -> dict | None:
    ds = str(raw["date"])
    d = date.fromisoformat(ds[:10]) if ds[:4].isdigit() else parse_date_he(ds)
    if d is None or not (C.PERIOD_START <= d <= C.PERIOD_END):
        return None
    signed = float(raw["amount"])
    if signed == 0:
        return None
    merchant = strip_rtl(raw.get("description") or "")
    acct = _account_kind(raw)
    currency = raw.get("currency", "ILS")
    tx = {
        "id": raw["id"], "owner": owner,
        "card": raw.get("accountId", ""),
        "source": "moneytor", "account_kind": acct,
        "tx_date": d.isoformat(), "bill_date": d.isoformat(),
        "merchant": merchant, "category": "",
        "moneytor_category": raw.get("category", ""), "type": raw.get("type", ""),
        "amount": abs(signed),
        "currency": "₪" if currency == "ILS" else currency,
        "orig_amount": abs(signed), "orig_currency": currency,
        "sub_type": "credit" if signed > 0 else "debit",
        "notes": raw.get("extra_info") or "",
        "foreign": currency != "ILS",
    }
    if acct == "bank":
        tx["bank_kind"] = _derive_bank_kind(merchant, signed, raw)
    return tx

# ─── Loaders ─────────────────────────────────────────────────────────────────

XLSX_HEADER_ROW = 4  # 1-indexed row with headers in Isracard exports

def load_xlsx_card(path: Path, owner: str) -> list[dict]:
    """Loads Isracard transaction export. Returns list of normalized tx dicts."""
    out = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5: continue
        # Header is row index 3 (4th)
        header = rows[3]
        col = {h: i for i, h in enumerate(header) if h}
        for r in rows[4:]:
            if not r or not r[0]: continue
            tx_date = r[col.get('תאריך עסקה')] if 'תאריך עסקה' in col else None
            merchant = strip_rtl(str(r[col.get('שם בית העסק', 1)] or ''))
            category = strip_rtl(str(r[col.get('קטגוריה', 2)] or ''))
            card = strip_rtl(str(r[col.get('4 ספרות אחרונות של כרטיס האשראי', 3)] or ''))
            sub_type = strip_rtl(str(r[col.get('סוג עסקה', 4)] or ''))
            amount = parse_amount(r[col.get('סכום חיוב', 5)])
            currency = strip_rtl(str(r[col.get('מטבע חיוב', 6)] or '₪'))
            orig_amt = parse_amount(r[col.get('סכום עסקה מקורי', 7)])
            orig_cur = strip_rtl(str(r[col.get('מטבע עסקה מקורי', 8)] or '₪'))
            bill_date = r[col.get('תאריך חיוב', 9)] if 'תאריך חיוב' in col else None
            notes = strip_rtl(str(r[col.get('הערות', 10)] or '')) if r[col.get('הערות', 10)] else ''

            d_tx = parse_date_he(str(tx_date)) if not isinstance(tx_date, (date, datetime)) else (tx_date.date() if isinstance(tx_date, datetime) else tx_date)
            d_bill = parse_date_he(str(bill_date)) if not isinstance(bill_date, (date, datetime)) else (bill_date.date() if isinstance(bill_date, datetime) else bill_date)

            # Filter: bill_date in window. EXCEPTION: utility-style merchants (electric/water/gas/arnona)
            # are bi-monthly in arrears — judge by tx_date so the consumption-period view is correct.
            UTIL_TX_RE = re.compile(r'חברת ה?חשמל|מקורות|תאגיד מים|מי [א-ת]+|חשמל\b|פז גז|סופר ?גז|אמישראגז|ארנונה|עיריית')
            in_bill_window = d_bill and PERIOD_START <= d_bill <= PERIOD_END
            in_tx_window_utility = d_tx and PERIOD_START <= d_tx <= PERIOD_END and UTIL_TX_RE.search(merchant)
            if not (in_bill_window or in_tx_window_utility):
                continue
            if amount == 0: continue  # skip 0 disputed lines
            out.append({
                'owner': owner,
                'card': card,
                'source': 'isracard-xlsx',
                'tx_date': d_tx.isoformat() if d_tx else '',
                'bill_date': d_bill.isoformat() if d_bill else '',
                'merchant': merchant,
                'category': category,
                'amount': amount,
                'currency': currency,
                'orig_amount': orig_amt,
                'orig_currency': orig_cur,
                'sub_type': sub_type,
                'notes': notes,
                'foreign': sn.startswith('עסקאות חו'),
            })
    return out

# ─── Isracard monthly PDF statement loader ──────────────────────────────────

# regex to match a transaction line in pdftotext -layout output (after stripping RTL marks)
# Lines look like:  AMOUNT  AMOUNT  category  card_type+merchant  DD/MM/YY
TX_LINE_RE = re.compile(
    r'^\s*([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+(.+?)\s+(\d{2}/\d{2}/\d{2})\s*$'
)
EXTRA_LINE_RE = re.compile(r'(תשלום\s+\d+\s+מתוך\s+\d+|הנחה.*?₪.*?\b)')

def load_isracard_pdf(path: Path, bill_year_month: tuple[int,int], owner: str, account: str = 'card') -> list[dict]:
    """Loads an Isracard monthly PDF statement. bill_year_month e.g. (2025, 10)."""
    txt = subprocess.check_output(
        ['pdftotext', '-layout', '-enc', 'UTF-8', str(path), '-'],
        stderr=subprocess.DEVNULL
    ).decode('utf-8')
    bill_date = date(bill_year_month[0], bill_year_month[1], 10)
    out = []
    in_foreign = False
    for raw in txt.splitlines():
        line = strip_rtl(raw)
        if 'עסקות בחו' in line or 'חו"ל ומט"ח' in line:
            in_foreign = True
        m = TX_LINE_RE.match(line)
        if not m: continue
        a1, a2, mid, ds = m.groups()
        # ds is DD/MM/YY
        try:
            tx_d = datetime.strptime(ds, '%d/%m/%y').date()
        except: continue
        # mid contains "category merchant" or "card_type merchant category" — split heuristically
        # The category is typically a short Hebrew tag with no digits, appearing right before amounts in original layout
        # After RTL/LTR layout fixing, structure is: AMOUNT AMOUNT CATEGORY MERCHANT_BLOB DATE
        # Take first 1-2 Hebrew tokens as category, rest as merchant
        mid_clean = re.sub(r'\s+', ' ', mid).strip()
        # Heuristic: category is the first token chunk (no digits, short, often containing '/')
        tokens = mid_clean.split(' ')
        # Walk from the left; collect category until token looks like merchant (e.g., contains a known card type)
        # Simpler: take first 1-2 tokens as category if they're known Hebrew labels
        known_cats = {
            'מכולת/סופר','דלק','מסעדות/קפה','מסעדות','שונות','אופנה','הלבשה','מעדניות',
            'תחבורה','קוסמטיקה','שירותי רכב','שירותי תקשורת','פנאי','פנאי, בידור וספורט',
            'בידור','ספורט','בריאות','חניונים','תחבורה ורכבים','קוסמטיקה וטיפוח',
            'מסעדות, קפה וברים','דלק/חנייה','אלקטרוניקה','תכשיטים','מתנות'
        }
        category = ''
        merchant_start = 0
        for n in (2, 1):
            cand = ' '.join(tokens[:n])
            if cand in known_cats:
                category = cand; merchant_start = n; break
        if not category:
            # default: first token
            category = tokens[0] if tokens else ''
            merchant_start = 1
        merchant = ' '.join(tokens[merchant_start:]).strip()
        # Strip card-type prefix
        merchant = re.sub(r'^(תש\.נייד|לא הוצג|תש\.רגיל|תש\.|רגילה)\s+', '', merchant)
        amount = parse_amount(a1)  # סכום החיוב (charged amount, post installments)
        out.append({
            'owner': owner,
            'card': account,
            'source': 'isracard-pdf',
            'tx_date': tx_d.isoformat(),
            'bill_date': bill_date.isoformat(),
            'merchant': merchant,
            'category': category,
            'amount': amount,
            'currency': '₪',
            'orig_amount': parse_amount(a2),
            'orig_currency': '₪',
            'sub_type': '',
            'notes': '',
            'foreign': in_foreign,
        })
    return out

# ─── Leumi bank PDF parser ────────────────────────────────────────────────────

# Description tokens to skip (card payments — already in card data) or note specially
BANK_SKIP_PATTERNS = [
    re.compile(r'^לאומי\s*ויזה'),  # Leumi Visa card payment
    re.compile(r'^מקס\s*איט'),     # Max It card payment
]
# RENT_AMOUNT (one person's rent share, transferred to settle a month) comes from config — see header.

# Couple-transfer detection (Bit/PayBox/digital transfer)
COUPLE_TRANSFER_RE = re.compile(r'הפועלים-ביט|פייבוקס|העברה ?דיגיט|^העברת')

def load_bank_discount_csv(path: Path, owner: str, account: str = 'bank') -> list[dict]:
    """Parses Discount Bank CSV export (UTF-16 LE, tab-delimited).
    Columns: תאריך | יום ערך | תיאור התנועה | זכות/חובה ₪ | יתרה ₪ | אסמכתה | עמלה | ערוץ
    """
    raw = path.read_bytes()
    if raw[:2] == b'\xff\xfe':
        text = raw.decode('utf-16-le')
    elif raw[:2] == b'\xfe\xff':
        text = raw.decode('utf-16-be')
    else:
        text = raw.decode('utf-8')
    text = text.lstrip('﻿')
    out = []
    for ln in text.splitlines()[1:]:
        if not ln.strip(): continue
        f = ln.split('\t')
        if len(f) < 5: continue
        d = parse_date_he(f[0])
        if not d or not (PERIOD_START <= d <= PERIOD_END): continue
        desc = f[2].strip()
        try: amt = float(f[3].strip().replace(',', ''))
        except: continue
        direction = 'credit' if amt > 0 else 'debit'
        amount = abs(amt)
        # Skip card-payment debits (already in card data)
        if 'מקס איט' in desc or 'ישראכרט' in desc: continue
        is_rent_xfer = abs(amount - RENT_AMOUNT) < 0.01 and ('שכר דירה' in desc or _has_token(desc, C.COUPLE_NAME_TOKENS))
        is_landlord = (amount == RENT_PER_MONTH_TOTAL and direction == 'debit' and ('שיק' in desc or _has_token(desc, C.LANDLORD_TOKENS)))
        is_salary = direction == 'credit' and _has_token(desc, C.SALARY_KEYWORDS)
        is_non_income = direction == 'credit' and (
            'הקמת הלוואה' in desc or ('שיק' in desc and 'עבר זמ' in desc)
            or _has_token(desc, C.SELF_TRANSFER_TOKENS)
        )
        bank_kind = 'other'
        if is_landlord: bank_kind = 'landlord_rent'
        elif is_rent_xfer: bank_kind = 'rent_settlement'
        elif is_salary: bank_kind = 'salary'
        elif is_non_income: bank_kind = 'non_income_credit'
        elif _has_token(desc, C.COUPLE_NAME_TOKENS): bank_kind = 'couple_transfer'
        out.append({
            'owner': owner, 'card': account, 'source': 'discount-bank',
            'tx_date': d.isoformat(), 'bill_date': d.isoformat(),
            'merchant': desc[:120], 'category': 'bank',
            'amount': amount, 'currency': '₪',
            'orig_amount': amount, 'orig_currency': '₪',
            'sub_type': direction, 'notes': '', 'foreign': False,
            'bank_kind': bank_kind,
        })
    return out


def load_bank_discount(path: Path, owner: str, account: str = 'bank') -> list[dict]:
    """Parses a Discount Bank PDF account statement. Returns bank-side movements."""
    txt = subprocess.check_output(
        ['pdftotext', '-layout', '-enc', 'UTF-8', str(path), '-'],
        stderr=subprocess.DEVNULL
    ).decode('utf-8')
    # Track current year/month from section headers
    HE_MONTHS = {'ינואר':1,'פברואר':2,'מרץ':3,'מרס':3,'אפריל':4,'מאי':5,'יוני':6,'יולי':7,'אוגוסט':8,'ספטמבר':9,'אוקטובר':10,'נובמבר':11,'דצמבר':12}
    cur_year = 2025; cur_month = 8
    out = []
    for raw in txt.splitlines():
        s = re.sub(r'[‎‏‪-‮⁦-⁩ ]', '', raw).strip()
        # Detect month section header (e.g. "אוקטובר2025")
        m = re.search(r'(ינואר|פברואר|מרץ|מרס|אפריל|מאי|יוני|יולי|אוגוסט|ספטמבר|אוקטובר|נובמבר|דצמבר)(\d{4})', s)
        if m:
            cur_month = HE_MONTHS[m.group(1)]
            cur_year  = int(m.group(2))
            continue
        # Skip non-transaction lines
        if not re.search(r'\d{2}/\d{2}', s): continue
        if 'תאריך' in s or 'ליום' in s and not re.search(r'^\d', s): continue
        # Find ALL dd/mm in the line; the LAST one before any trailing description is the posting day.
        # Some rows put description at the very end, after the dates.
        date_matches = list(re.finditer(r'(\d{2})/(\d{2})(?!\d)', s))
        if not date_matches: continue
        last_date_match = date_matches[-1]
        try:
            d = date(cur_year, cur_month, int(last_date_match.group(1)))
        except ValueError:
            continue
        if not (PERIOD_START <= d <= PERIOD_END):
            continue
        # body = everything before the final block of dates
        # If there are 2 adjacent dates (value + posting), strip both
        body = s[:last_date_match.start()]
        # Trailing description (after the dates) — capture and append to body
        trailing = s[last_date_match.end():]
        # Strip any preceding dd/mm (value date) at end of body
        m_val = re.search(r'(\d{2}/\d{2})$', body)
        if m_val: body = body[:m_val.start()]
        # Combine body + trailing into "data + description"
        body = body + ' ' + trailing
        # Find all signed/unsigned numbers in the body
        nums = re.findall(r'-?[\d,]+\.\d{2}', body)
        if not nums: continue
        # Heuristic: in pdftotext layout the columns appear in this order (after RTL flip):
        # balance, debit (negative), credit (positive), reference#, description, value-date, day
        # When stripped, the FIRST number in body = balance (or sometimes debit if balance carries through).
        # The number that's the transaction amount is typically the one near the description.
        # Simpler approach: take LAST number as the amount (it's adjacent to the asmachta/desc).
        # Direction: check if number is negative (debit) or positive (credit).
        amount_str = nums[-1]
        amount = abs(parse_amount(amount_str))
        direction = 'debit' if amount_str.startswith('-') else 'credit'
        # Description: text between last number and any asmachta digits
        desc_part = body[body.rfind(amount_str)+len(amount_str):]
        # Strip leading asmachta (digits, optional slash+digits, optional time hh:mm)
        desc_part = re.sub(r'^[\d/]+(?:\d{2}:\d{2})?', '', desc_part).strip()
        merchant = desc_part if desc_part else 'discount-bank-tx'
        # Categorize
        bank_kind = 'other'
        is_landlord_cheque = (amount == RENT_PER_MONTH_TOTAL
            and direction == 'debit'
            and ('שיק' in merchant or 'משיכתשיק' in merchant.replace(' ','') or _has_token(merchant, C.LANDLORD_TOKENS)))
        is_couple_xfer = _has_token(merchant, C.COUPLE_NAME_TOKENS)
        is_salary = _has_token(merchant, C.SALARY_KEYWORDS) and direction == 'credit'
        # Non-income credits: loan setups, returned cheques, intra-self transfers
        is_non_income_credit = direction == 'credit' and (
            'הקמתהלוואה' in merchant or 'הקמת הלוואה' in merchant
            or ('שיק' in merchant and 'עברזמ' in merchant)  # returned cheque
            or ('שיק' in merchant and 'עבר זמ' in merchant)
            or _has_token(merchant, C.SELF_TRANSFER_TOKENS)  # intra-self
        )
        is_card_payment = ('ישראכרט' in merchant or 'מקסאיט' in merchant or 'מקס איט' in merchant)
        if is_card_payment:
            continue  # skip — already in card data
        if is_landlord_cheque:
            bank_kind = 'landlord_rent'
        elif is_couple_xfer:
            bank_kind = 'rent_settlement' if abs(amount-RENT_AMOUNT)<0.01 else 'couple_transfer'
        elif is_salary:
            bank_kind = 'salary'
        elif is_non_income_credit:
            bank_kind = 'non_income_credit'  # exclude from earnings
        out.append({
            'owner': owner,
            'card': account,
            'source': 'discount-bank',
            'tx_date': d.isoformat(),
            'bill_date': d.isoformat(),
            'merchant': merchant[:80],
            'category': 'bank',
            'amount': amount,
            'currency': '₪',
            'orig_amount': amount,
            'orig_currency': '₪',
            'sub_type': direction,
            'notes': '',
            'foreign': False,
            'bank_kind': bank_kind,
        })
    return out


def load_bank_leumi(path: Path, owner: str, account: str = 'bank') -> list[dict]:
    """Parses a Bank Leumi PDF statement. Returns bank-side movements (skipping card payments)."""
    txt = subprocess.check_output(
        ['pdftotext', '-layout', '-enc', 'UTF-8', str(path), '-'],
        stderr=subprocess.DEVNULL
    ).decode('utf-8')
    raws = []
    for line in txt.splitlines():
        s = re.sub(r'[‎‏‪-‮⁦-⁩ ]', '', line).strip()
        # Pattern: <balance><amount><asmachta><desc><date><date>
        m = re.match(r'^([\d,]+\.\d{2})([\d,]+\.\d{2})(\d+)(.+?)(\d{2}/\d{2}/\d{4})(\d{2}/\d{2}/\d{4})$', s)
        if not m: continue
        bal = parse_amount(m.group(1))
        amt = parse_amount(m.group(2))
        desc = m.group(4).strip()
        d = parse_date_he(m.group(5))
        raws.append({'date': d, 'balance': bal, 'amount': amt, 'desc': desc})
    # Sort oldest → newest and compute direction by balance change
    raws.sort(key=lambda r: r['date'])
    prev_bal = None
    out = []
    for r in raws:
        if not (PERIOD_START <= r['date'] <= PERIOD_END):
            prev_bal = r['balance']
            continue
        direction = None
        if prev_bal is not None:
            delta = r['balance'] - prev_bal
            direction = 'credit' if delta > 0 else 'debit'
        prev_bal = r['balance']
        # Skip credit card payments (already in card data)
        if any(p.search(r['desc']) for p in BANK_SKIP_PATTERNS):
            continue
        # Categorize
        is_transfer = bool(COUPLE_TRANSFER_RE.search(r['desc']))
        is_rent = is_transfer and abs(r['amount'] - RENT_AMOUNT) < 0.01
        # Landlord rent: ₪5,400 check (₪5,400 שיק)
        is_landlord = (r['desc'] == 'שיק' and abs(r['amount'] - RENT_PER_MONTH_TOTAL) < 0.01)
        out.append({
            'owner': owner,
            'card': account,
            'source': 'leumi-bank',
            'tx_date': r['date'].isoformat(),
            'bill_date': r['date'].isoformat(),
            'merchant': r['desc'],
            'category': 'bank',
            'amount': r['amount'],
            'currency': '₪',
            'orig_amount': r['amount'],
            'orig_currency': '₪',
            'sub_type': direction or '',
            'notes': '',
            'foreign': False,
            'bank_kind': ('landlord_rent' if is_landlord else ('rent_settlement' if is_rent else ('couple_transfer' if is_transfer else 'other'))),
        })
    return out

# ─── Load everything ──────────────────────────────────────────────────────────

def _ym_from_name(fp: str) -> tuple[int, int]:
    """Bill month from a file named YYYY-MM.pdf."""
    y, m = map(int, Path(fp).stem.split('-'))
    return (y, m)


# Maps a Source.kind (config) → a loader call. Add an entry here when you add a parser.
LOADERS = {
    'isracard_xlsx':        lambda fp, owner, acct: load_xlsx_card(fp, owner),
    'isracard_pdf_monthly': lambda fp, owner, acct: load_isracard_pdf(fp, _ym_from_name(fp), owner, acct or 'card'),
    'leumi_pdf':            lambda fp, owner, acct: load_bank_leumi(fp, owner, acct or 'bank'),
    'discount_csv':         lambda fp, owner, acct: load_bank_discount_csv(fp, owner, acct or 'bank'),
    'discount_pdf':         lambda fp, owner, acct: load_bank_discount(fp, owner, acct or 'bank'),
}


def load_all() -> list[dict]:
    """Walks config.PEOPLE → their sources → the matching parser."""
    txs = []
    for person in C.PEOPLE:
        base = ROOT / person.data_dir
        # A Discount account given as both CSV and PDF: CSV wins, PDF fills gaps.
        # This dedupe set is per-person so each account is independent.
        discount_seen = set()
        for src in person.sources:
            loader = LOADERS.get(src.kind)
            if loader is None:
                print(f'WARN: unknown source kind {src.kind!r} for {person.id}', file=sys.stderr)
                continue
            files = sorted(glob.glob(str(base / src.glob)))
            if not files:
                print(f'WARN: no files matched {person.id}/{src.glob}', file=sys.stderr)
            for fp in files:
                rows = loader(Path(fp), person.id, src.account)
                if src.kind == 'discount_csv':
                    for t in rows:
                        key = (t['tx_date'], round(t['amount'], 2), t['merchant'][:30])
                        if key in discount_seen: continue
                        discount_seen.add(key); txs.append(t)
                elif src.kind == 'discount_pdf':
                    for t in rows:  # approximate dedupe — PDF descriptions are mangled vs CSV
                        key = (t['tx_date'], round(t['amount'], 2))
                        if any(k[:2] == key for k in discount_seen): continue
                        discount_seen.add(key + (t['merchant'][:30],))
                        txs.append(t)
                else:
                    txs += rows
    return txs

# ─── Classification ──────────────────────────────────────────────────────────

# Merchant substring → tag. First match wins.
UTILITY_RULES = [
    # (regex, tag)
    (re.compile(r'חברת ה?חשמל|^חשמל\b'), 'utility:electric'),
    (re.compile(r'מקורות|תאגיד מים|מי [א-ת]+\b|הגיחון|מי רעננה|מי שקמה|מי אביבים|פלגי השרון'), 'utility:water'),
    (re.compile(r'פז גז|סופר ?גז|אמישראגז|דור גז|גז המגן|פזגז|אמיש?רגז'), 'utility:gas'),
    (re.compile(r'בזק\b|הוט\b|HOT\b|YES\b|יס\b|נטוויז|טריפל|triple', re.I), 'utility:internet'),
    (re.compile(r'ארנונה|עירייה|עיריית'), 'utility:arnona'),
    (re.compile(r'ועד בית|וועד בית|ועד הבית'), 'utility:building'),
    (re.compile(r'משכנתא|שכר דירה|שכ"ד|שכ״ד'), 'utility:rent_mortgage'),
    (re.compile(r'ביטוח לאומי|בטוח לאומי'), 'utility:bituach_leumi'),
]
# Insurance & vehicle = personal per user rule, classified by category lists above

TRANSFER_CATEGORIES = {'העברת כספים', 'משיכת מזומן'}
TRANSFER_MERCHANT_PATTERNS = [
    re.compile(r'^BIT$|העברה ב BIT|PAYBOX|דמי כרטיס|כספומט', re.I),
]
FUEL_MERCHANT_RE = re.compile(r'\bפז\b|דור אלון|סונול|דלק מנטה|רות\b|סד"ש|אברך|ספרינט-?אש|פז יילו|אפליקציית יילו|דלק קמעונאות')
PHONE_MERCHANT_RE = re.compile(r'סלקום|פרטנר|partner|cellcom|012|019|פלאפון|רמי לוי תקשורת', re.I)

SHARED_CATEGORIES = {
    'מכולת/סופר',  # grocery
    'סופר',
    'מכולת',
    'מזון וצריכה',  # Person A's xlsx label for groceries+consumables
    'חיות מחמד',  # pet
    # עיצוב הבית / בית וגן moved to FURNITURE_CATEGORIES (own section)
}

PERSONAL_CATEGORIES = {
    'אופנה', 'הלבשה',  # fashion
    'קוסמטיקה', 'קוסמטיקה וטיפוח',
    'תכשיטים',
    'רפואה ובתי מרקחת',  # medical/pharmacy — usually personal
    'פארמה',
    'חשמל ומחשבים',  # electronics/computers — usually personal
    'ספרים ודפוס',  # books — usually personal
    'מכוני',  # gyms/studios
    'ביטוח',  # insurance — user rule: not split
    'תחבורה ורכבים',  # vehicles — user rule: not split (incl insurance/maintenance)
    'שירותי רכב',  # car services
    'חניונים',  # parking
    'עירייה וממשלה',  # gov/DMV — user rule: vehicle related, personal
}

# Gym/fitness merchants — personal
GYM_MERCHANT_RE = re.compile(r'O-LIVE|פיזיקל|פילאטיס|יוגה|אנרגיה רעננה|עידן אלפיים|holmes|GYM\b|fitness|crossfit|פעמיים בשבוע', re.I)
# Subscription services — personal (entertainment)
SUBSCRIPTION_MERCHANT_RE = re.compile(r'NETFLIX|APPLE\.COM|YOUTUBEPREMIUM|SPOTIFY|DISNEY|HBO|REAL-DEBRID', re.I)
# Business / SaaS / dev-tool subscriptions — personal but tagged business (Person A)
BUSINESS_MERCHANT_RE = re.compile(r'CLAUDE\.AI|ANTHROPIC|AWS\s|AMAZON ?WEB ?SERV|GITHUB|VERCEL|OPENAI|GOOGLE CLOUD|DIGITALOCEAN|HEROKU|LINEAR|STRIPE|PADDLE|FIGMA|NOTION|POSTMAN|SENTRY|SUPABASE|RENDER\.COM|RAILWAY|CURSOR|REPLIT|COPILOT|POSTHOG|TYPEFORM|MIDJOURNEY|CLOUDFLARE|OPENROUTER|GOOGLE ONE|GOOGLE\*GOOGLE|HEADSTART|PDFGURU', re.I)
# Vehicle/DMV
VEHICLE_MERCHANT_RE = re.compile(r'משרד התחבורה|רשיונ|רישוי|מוסך|טסט|איתוראן|איתורן|מיגון|פנגו|כביש 6|דרך ארץ|cello|מאסטר רכב|רכבים', re.I)

RESTAURANT_CATEGORIES = {'מסעדות', 'מסעדות/קפה', 'מסעדות, קפה וברים', 'מעדניות'}
VACATION_CATEGORIES = {'טיסות ותיירות'}
FURNITURE_CATEGORIES = {'עיצוב הבית', 'בית וגן', 'ריהוט'}
FURNITURE_MERCHANT_RE = re.compile(r'איקאה|IKEA|הום סנטר|home ?center|פוקס הום|fox home|אשלי|ashley|פנדה הום|panda home|רהיט|ריהוט|take it|הביתה|home essentials', re.I)
# Merchant patterns for vacations (since some travel hides under other categories)
VACATION_MERCHANT_RE = re.compile(r'AIRBNB|BOOKING|HOTEL|מלון|מלונות|נופש|airbnb|booking|hotel', re.I)

AMBIGUOUS_CATEGORIES = {
    'שונות',
    'פנאי', 'פנאי, בידור וספורט', 'בידור', 'ספורט',
    'תחבורה', 'דלק', 'דלק/חנייה',  # vehicles already personal above; these stay flagged
    'דלק, חשמל וגז',  # composite — split by merchant
    'בריאות',
    'מתנות',
    'אלקטרוניקה',
    'שירותי תקשורת',  # could be personal phone or home internet
    'ביגוד והנעלה',
    'עירייה וממשלה',
}

# Merchant patterns that override category
SHARED_MERCHANT_PATTERNS = [
    re.compile(r'סופרמרקט|מכולת|שופרסל|רמי לוי|יוחננוף|ויקטורי|אושר עד|חצי חינם|מגה|טיב טעם|carrefour|CARREFOUR|סופר ?פארם|super-pharm|פוקס הום|home center|הום סנטר|איקאה|IKEA', re.I),
    re.compile(r'גוד פארם|good pharm|בי פארם|פארמה|pharm', re.I),
    re.compile(r'בינז קפה|beanz|beans coffee', re.I),  # house coffee subscription
]

def classify_deterministic(tx: dict):
    """Returns tx with added: tag, status (shared|personal|flag|transfer|rent), reason.
    Returns None if the cascade cannot decide (residue — escalate to LLM harness)."""
    merchant = tx['merchant']
    category = tx['category']
    if tx.get('account_kind') == 'card' and tx.get('sub_type') == 'credit':
        return {**tx, 'tag': 'refund', 'status': 'refund', 'reason': 'card credit / refund — excluded from spend'}
    # Bank-side transactions — routed to dedicated buckets (kept out of flag/shared/personal)
    if tx.get('account_kind') == 'bank':
        kind = tx.get('bank_kind')
        direction = tx.get('sub_type')
        if kind == 'landlord_rent':
            return {**tx, 'tag': 'landlord_rent', 'status': 'landlord_rent', 'reason': '₪5,400 = rent paid to landlord'}
        if kind == 'rent_settlement':
            return {**tx, 'tag': 'rent_settlement', 'status': 'rent', 'reason': '2,700 transfer = rent settlement'}
        if kind == 'salary':
            return {**tx, 'tag': 'salary', 'status': 'earnings', 'reason': 'salary credit', 'earnings_kind': 'salary'}
        # Loan repayments and loan setups (BEFORE non_income_credit so the setup is captured here)
        if 'פירעון הלוואה' in merchant or 'פירעוןהלוואה' in merchant:
            # Extract loan ID (last number sequence)
            m = re.search(r'(\d{8,})', merchant.replace(' ', ''))
            loan_id = m.group(1) if m else 'unknown'
            return {**tx, 'tag': 'loan_payment', 'status': 'loan', 'reason': f'loan repayment {loan_id}', 'loan_id': loan_id}
        if 'הקמת הלוואה' in merchant or 'הקמתהלוואה' in merchant:
            m = re.search(r'(\d{8,})', merchant.replace(' ', ''))
            loan_id = m.group(1) if m else 'unknown'
            return {**tx, 'tag': 'loan_received', 'status': 'loan', 'reason': f'loan setup {loan_id}', 'loan_id': loan_id}
        if kind == 'non_income_credit':
            return {**tx, 'tag': 'non_income_credit', 'status': 'transfer', 'reason': 'returned cheque / intra-self — not income'}
        # Rental income — ₪2,100 credits from Person A's apartment (must come BEFORE generic earnings catch)
        if direction == 'credit' and abs(tx['amount'] - RENTAL_INCOME_AMOUNT) < 0.01:
            return {**tx, 'tag': 'rental_income', 'status': 'rental_income', 'reason': '₪2,100 credit = apartment rental income'}
        # Couple transfers: only if direction is debit (outgoing). Credit-direction "transfer" is
        # most likely incoming salary/payment, route to earnings.
        if kind == 'couple_transfer' and direction == 'debit':
            return {**tx, 'tag': 'couple_transfer', 'status': 'rent' if _has_token(merchant, C.COUPLE_NAME_TOKENS) else 'transfer', 'reason': 'transfer out'}
        # All other credits → earnings (salary or other income)
        if direction == 'credit':
            return {**tx, 'tag': 'earnings', 'status': 'earnings', 'reason': 'bank credit'}
        # Debits routing
        if 'עמל' in merchant:
            return {**tx, 'tag': 'bank_fee', 'status': 'personal', 'reason': 'bank fee'}
        if 'משיכה' in merchant or 'המרת' in merchant or 'עמלת' in merchant:
            return {**tx, 'tag': 'cash_or_fx', 'status': 'transfer', 'reason': 'cash withdrawal / fx'}
        if kind == 'couple_transfer':
            return {**tx, 'tag': 'bank_transfer_out', 'status': 'transfer', 'reason': 'bank transfer out (Bit/PayBox/digital)'}
        return {**tx, 'tag': 'bank_other', 'status': 'transfer', 'reason': 'bank debit — unclassified'}
    # Card-side rent settlement: a one-share BIT/PayBox transfer from person_a's card = rent paid to person_b
    if RENT_AMOUNT and abs(tx['amount'] - RENT_AMOUNT) < 0.01 and ('BIT' in tx.get('notes','').upper() or 'PAYBOX' in tx.get('notes','').upper() or 'BIT' in merchant.upper() or 'PAYBOX' in merchant.upper()):
        return {**tx, 'tag': 'rent_settlement', 'status': 'rent',
                'reason': f'{CURRENCY}{RENT_AMOUNT:,.0f} BIT/transfer from card = rent settlement',
                'sub_type': 'debit' if tx.get('owner')=='person_a' else 'credit'}
    # Transfer / cash withdrawal — exclude from spend
    for pat in TRANSFER_MERCHANT_PATTERNS:
        if pat.search(merchant):
            return {**tx, 'tag': 'transfer', 'status': 'transfer', 'reason': 'transfer merchant'}
    # Utility check
    for pat, tag in UTILITY_RULES:
        if pat.search(merchant):
            return {**tx, 'tag': tag, 'status': 'shared', 'reason': f'utility match: {tag}'}
    # Fuel — personal (each their own car)
    if FUEL_MERCHANT_RE.search(merchant):
        return {**tx, 'tag': 'fuel', 'status': 'personal', 'reason': 'fuel merchant'}
    # Phone — personal per user rule
    if PHONE_MERCHANT_RE.search(merchant):
        return {**tx, 'tag': 'phone', 'status': 'personal', 'reason': 'phone bill (user rule)'}
    # Gym/fitness — personal
    if GYM_MERCHANT_RE.search(merchant):
        return {**tx, 'tag': 'gym', 'status': 'personal', 'reason': 'gym/fitness merchant'}
    # Business / SaaS / dev-tools (must match BEFORE generic subscription rule)
    if BUSINESS_MERCHANT_RE.search(merchant):
        return {**tx, 'tag': 'business', 'status': 'business', 'reason': 'business / SaaS / dev tool'}
    # Subscriptions (Netflix/Apple/YouTube etc.) — personal entertainment
    if SUBSCRIPTION_MERCHANT_RE.search(merchant):
        return {**tx, 'tag': 'subscription', 'status': 'personal', 'reason': 'subscription service'}
    # Vehicle merchants — personal per user rule (DMV, mechanic, parking app, toll roads)
    if VEHICLE_MERCHANT_RE.search(merchant):
        return {**tx, 'tag': 'vehicle', 'status': 'personal', 'reason': 'vehicle merchant'}
    # Furniture / home decor — shared, but tagged for its own section
    if FURNITURE_MERCHANT_RE.search(merchant):
        return {**tx, 'tag': 'furniture', 'status': 'shared', 'reason': 'furniture / home decor'}
    # Shared merchant patterns (matches BEFORE restaurant so e.g. בינז קפה / IKEA win)
    for pat in SHARED_MERCHANT_PATTERNS:
        if pat.search(merchant):
            return {**tx, 'tag': 'grocery_or_household', 'status': 'shared', 'reason': 'shared merchant'}
    # Vacations — own bucket, not distributed
    if VACATION_MERCHANT_RE.search(merchant):
        return {**tx, 'tag': 'vacation', 'status': 'vacation', 'reason': 'vacation (travel/hotel/flight)'}
    return None  # residue — re-grounded cascade can't decide; escalate to the LLM harness

TAG_LABELS = {
    'grocery_or_household': 'מכולת/בית', 'grocery': 'מכולת', 'fuel': 'דלק', 'phone': 'תקשורת',
    'gym': 'כושר', 'business': 'עסקי', 'subscription': 'מנויים', 'vehicle': 'רכב',
    'furniture': 'ריהוט', 'restaurant': 'מסעדות', 'vacation': 'חופשה', 'refund': 'זיכוי',
    'transfer': 'העברה', 'salary': 'משכורת', 'earnings': 'הכנסה', 'landlord_rent': 'שכר דירה',
    'rent_settlement': 'התחשבנות שכ״ד', 'couple_transfer': 'העברה זוגית',
    'non_income_credit': 'זיכוי לא-הכנסה', 'loan_payment': 'החזר הלוואה',
    'loan_received': 'קבלת הלוואה', 'rental_income': 'הכנסה משכירות', 'bank_fee': 'עמלת בנק',
    'cash_or_fx': 'מזומן/מט״ח', 'bank_transfer_out': 'העברה יוצאת', 'bank_other': 'בנק-אחר',
    'utility:electric': 'חשמל', 'utility:water': 'מים', 'utility:gas': 'גז',
    'utility:internet': 'אינטרנט', 'utility:arnona': 'ארנונה', 'utility:building': 'ועד בית',
    'utility:rent_mortgage': 'משכנתא', 'utility:bituach_leumi': 'ביטוח לאומי',
    'review': 'לבדיקה', 'foreign': 'חו״ל', 'other': 'אחר',
}

def display_category(tag: str) -> str:
    return TAG_LABELS.get(tag, tag)

def classify_one(tx, store, queue, *, rubric_hash, retune=False):
    rec = _store.lookup(store, tx, rubric_hash=rubric_hash, retune=retune)
    if rec is not None:
        return {**tx, 'tag': rec['tag'], 'status': rec['status'],
                'reason': rec['reason'], 'decided_by': rec['decided_by'], **rec.get('extra', {})}
    decided = classify_deterministic(tx)
    if decided is not None:
        _store.put_decision(store, tx, status=decided['status'], tag=decided['tag'],
                            reason=decided['reason'], decided_by='rule', rubric_hash=rubric_hash)
        return {**decided, 'decided_by': 'rule'}
    queue.append(tx)
    return None

def classify_all(txs, store, queue, *, rubric_hash, retune=False, allow_llm=True):
    cls = []
    for tx in txs:
        out = classify_one(tx, store, queue, rubric_hash=rubric_hash, retune=retune)
        if out is not None:
            cls.append(out)
    return cls

# ─── Aggregations ────────────────────────────────────────────────────────────

def month_key(iso: str) -> str:
    return iso[:7]

def main():
    txs = load_all()
    print(f'Loaded {len(txs)} transactions', file=sys.stderr)
    cls = [r for r in (classify_deterministic(t) for t in txs) if r is not None]  # stopgap (Task 8 replaces with store-backed classify_all)

    # Write CSV for inspection
    all_fields = sorted({k for t in cls for k in t.keys()})
    with open(OUTDIR/'transactions.csv', 'w', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=all_fields, extrasaction='ignore')
        w.writeheader()
        for t in cls:
            w.writerow({k: t.get(k, '') for k in all_fields})

    # Aggregations
    totals = {'person_a': 0.0, 'person_b': 0.0}
    by_status = defaultdict(lambda: {'person_a':0.0,'person_b':0.0})
    by_tag = defaultdict(lambda: {'person_a':0.0,'person_b':0.0})
    by_month = defaultdict(lambda: defaultdict(lambda: {'person_a':0.0,'person_b':0.0}))
    utilities = defaultdict(lambda: {'person_a':0.0,'person_b':0.0, 'items': []})
    flagged = []
    shared_items = []
    personal_items = defaultdict(list)

    transfers = []
    rent_settlements = []
    bank_credits = []     # for earnings section
    bank_debits = []
    restaurants = []
    vacations = []
    rental_incomes = []
    landlord_payments = []  # ₪5,400 checks (Person A paying landlord)
    for t in cls:
        if t.get('source') in ('leumi-bank', 'discount-bank') and t.get('sub_type') == 'credit':
            bank_credits.append(t)
        elif t.get('source') in ('leumi-bank', 'discount-bank') and t.get('sub_type') == 'debit':
            bank_debits.append(t)
        if t['status'] == 'restaurant':
            restaurants.append(t)
        elif t['status'] == 'vacation':
            vacations.append(t)
        elif t['status'] == 'rental_income':
            rental_incomes.append(t)
        elif t['status'] == 'landlord_rent':
            landlord_payments.append(t)
        o = t['owner']; a = t['amount']
        by_status[t['status']][o] += a
        m = month_key(t['bill_date'])
        by_month[m][t['status']][o] += a
        if t['status'] == 'transfer':
            transfers.append(t); continue
        if t['status'] == 'rent':
            rent_settlements.append(t); continue
        if t['status'] in ('restaurant', 'vacation', 'rental_income', 'earnings', 'landlord_rent', 'business', 'loan'):
            # tracked separately; not in cards spending totals (landlord rent handled in rent math below)
            continue
        totals[o] += a
        by_tag[t['tag']][o] += a
        if t['tag'].startswith('utility:'):
            utilities[t['tag']][o] += a
            utilities[t['tag']]['items'].append(t)
        if t['status'] == 'flag':
            flagged.append(t)
        elif t['status'] == 'shared':
            shared_items.append(t)
        else:
            personal_items[o].append(t)

    # Cards-only settlement math (50/50)
    shared_cards_person_a = by_status['shared']['person_a']
    shared_cards_person_b = by_status['shared']['person_b']
    shared_cards_total = shared_cards_person_a + shared_cards_person_b
    delta_cards = shared_cards_person_a - shared_cards_total/2

    # Rent math (now data-driven from -₪5,400 cheques on Person A's Leumi)
    assumed_rent_total = RENT_PER_MONTH_TOTAL * MONTHS_COUNT
    assumed_rent_each  = assumed_rent_total * RENT_SPLIT
    # Months Person A paid landlord (from his bank cheques)
    # Dedupe by (owner, year-month) — multiple cheque events in same month = one rent payment
    seen_owner_month = set()
    dedup_landlord = []
    for t in sorted(landlord_payments, key=lambda x: x['bill_date']):
        key = (t['owner'], t['bill_date'][:7])
        if key in seen_owner_month: continue
        seen_owner_month.add(key)
        dedup_landlord.append(t)
    landlord_payments = dedup_landlord
    person_a_landlord_items = [t for t in landlord_payments if t['owner']=='person_a']
    person_b_landlord_items = [t for t in landlord_payments if t['owner']=='person_b']
    person_a_landlord_months = sorted({t['bill_date'][:7] for t in person_a_landlord_items})
    person_b_landlord_months = sorted({t['bill_date'][:7] for t in person_b_landlord_items})
    person_a_landlord_count  = len(person_a_landlord_months)
    person_b_landlord_count_visible = len(person_b_landlord_months)
    # If we can see Person B's bank, count her months directly; otherwise infer from gap
    if person_b_landlord_count_visible > 0:
        person_b_landlord_count = person_b_landlord_count_visible
        person_b_landlord_paid_inferred = person_b_landlord_count * RENT_PER_MONTH_TOTAL
    else:
        person_b_landlord_count = MONTHS_COUNT - person_a_landlord_count
        person_b_landlord_paid_inferred = person_b_landlord_count * RENT_PER_MONTH_TOTAL
    person_a_landlord_paid   = person_a_landlord_count * RENT_PER_MONTH_TOTAL
    # Each owes half of total
    rent_person_a_owes_share  = assumed_rent_each
    rent_person_b_owes_share = assumed_rent_each
    # Net rent debt before settlement transfers: person_a paid - person_a's share
    rent_net_person_a = person_a_landlord_paid - rent_person_a_owes_share  # positive = person_a overpaid → Person B owes Person A
    # Observed rent settlements. Same transfer appears on both sides — dedupe by (date, amount, direction).
    # Direction normalized: + if Person A→Person B, − if Person B→Person A.
    unique_settle = {}
    for t in rent_settlements:
        if t['owner']=='person_a':
            sign = +1 if t.get('sub_type')=='debit' else -1
        else:  # person_b
            sign = +1 if t.get('sub_type')=='credit' else -1
        key = (t['tx_date'], round(t['amount'],2), sign)
        if key in unique_settle: continue
        unique_settle[key] = t
    rent_settlements = list(unique_settle.values())
    rent_person_a_to_person_b = sum(round(t['amount'],2) for k,t in unique_settle.items() if k[2]==+1)
    rent_person_b_to_person_a = sum(round(t['amount'],2) for k,t in unique_settle.items() if k[2]==-1)
    # Strict math (only counting what we see in Person A's data within window):
    # Person A's total contribution = (paid to landlord) + (sent to Person B as share) - (received from Person B as share)
    person_a_rent_contribution = person_a_landlord_paid + rent_person_a_to_person_b - rent_person_b_to_person_a
    rent_strict_owed_to_person_a = person_a_rent_contribution - rent_person_a_owes_share
    # Practical assumption: rent counter-transfers from Person B's bank are invisible to us,
    # and Person A's settlements occasionally land a few days outside the window.
    # User's view: every month gets settled somehow → rent nets to 0 over the long run.
    rent_remaining_owed_to_person_a = 0.0

    # Shared totals including rent (informational)
    shared_person_a_total = shared_cards_person_a + assumed_rent_each
    shared_person_b_total = shared_cards_person_b + assumed_rent_each
    shared_total_with_rent = shared_person_a_total + shared_person_b_total
    each_should_pay = shared_total_with_rent / 2
    # Final delta: cards + rent remainder
    final_delta = delta_cards + rent_remaining_owed_to_person_a

    # Earnings (Person A from Leumi, Person B from Discount)
    for t in bank_credits:
        if t.get('tag') == 'rental_income':
            t['earnings_kind'] = 'rental'
        elif t.get('tag') == 'salary' or (t['amount'] >= SALARY_MIN_AMOUNT and bool(SALARY_MERCHANT_RE.search(t['merchant']))):
            t['earnings_kind'] = 'salary'
            t['is_likely_salary'] = True
        else:
            t['earnings_kind'] = 'other'
    person_a_credits = [t for t in bank_credits if t['owner']=='person_a']
    person_b_credits = [t for t in bank_credits if t['owner']=='person_b']
    earnings_person_a_total  = sum(t['amount'] for t in person_a_credits)
    earnings_person_a_salary = sum(t['amount'] for t in person_a_credits if t.get('earnings_kind')=='salary')
    earnings_person_a_rental = sum(t['amount'] for t in person_a_credits if t.get('earnings_kind')=='rental')
    earnings_person_a_other  = earnings_person_a_total - earnings_person_a_salary - earnings_person_a_rental
    earnings_person_b_total  = sum(t['amount'] for t in person_b_credits)
    earnings_person_b_salary = sum(t['amount'] for t in person_b_credits if t.get('earnings_kind')=='salary')
    earnings_person_b_other  = earnings_person_b_total - earnings_person_b_salary

    # Restaurant + Vacation totals (just informational)
    rest_total = sum(t['amount'] for t in restaurants)
    rest_by_owner = {o: sum(t['amount'] for t in restaurants if t['owner']==o) for o in ('person_a','person_b')}
    vac_total = sum(t['amount'] for t in vacations)
    vac_by_owner = {o: sum(t['amount'] for t in vacations if t['owner']==o) for o in ('person_a','person_b')}

    # Per-car breakdown (vehicle/fuel/insurance etc. by owner)
    CAR_CATEGORIES = {'תחבורה ורכבים','שירותי רכב','חניונים','עירייה וממשלה','ביטוח'}
    CAR_TAGS = {'fuel','vehicle','phone'}  # phone shouldn't really be here but tagged similarly
    cars = {'person_a': [], 'person_b': []}
    for t in cls:
        if t['status'] != 'personal': continue
        if t['tag'] in ('fuel','vehicle'):
            cars[t['owner']].append(t)
        elif t['category'] in CAR_CATEGORIES and t['tag'] not in ('phone','gym','subscription'):
            cars[t['owner']].append(t)
    car_totals = {o: sum(t['amount'] for t in cars[o]) for o in cars}
    # Break down by sub-tag
    def car_breakdown(items):
        bd = defaultdict(float)
        for t in items:
            key = 'Fuel' if t['tag']=='fuel' else 'Vehicle / DMV / Maintenance' if t['tag']=='vehicle' else 'Insurance' if 'ביטוח' in t['category'] else 'Other car'
            bd[key] += t['amount']
        return dict(bd)
    car_breakdown_r = car_breakdown(cars['person_a'])
    car_breakdown_a = car_breakdown(cars['person_b'])

    # Subscriptions + Gym totals (informational)
    subs = [t for t in cls if t.get('tag')=='subscription']
    gyms = [t for t in cls if t.get('tag')=='gym']
    subs_total = sum(t['amount'] for t in subs)
    gyms_total = sum(t['amount'] for t in gyms)

    # Furniture / Home (shared but shown as own section)
    furniture = [t for t in cls if t.get('tag')=='furniture']
    furniture_total = sum(t['amount'] for t in furniture)
    furniture_by_owner = {o: sum(t['amount'] for t in furniture if t['owner']==o) for o in ('person_a','person_b')}

    # Business expenses (Person A — SaaS/dev/AI tools, not split with Person B)
    business = [t for t in cls if t.get('status')=='business']
    business_total = sum(t['amount'] for t in business)
    business_by_owner = {o: sum(t['amount'] for t in business if t['owner']==o) for o in ('person_a','person_b')}

    # Loans — aggregate by loan_id, separate repayments from setups
    loans_raw = [t for t in cls if t.get('status')=='loan']
    loans_by_id = defaultdict(lambda: {'owner': '', 'repayments': [], 'setups': []})
    for t in loans_raw:
        lid = t.get('loan_id', 'unknown')
        loans_by_id[lid]['owner'] = t['owner']
        if t.get('tag') == 'loan_received':
            loans_by_id[lid]['setups'].append(t)
        else:
            loans_by_id[lid]['repayments'].append(t)
    loans_summary = []
    for lid, info in loans_by_id.items():
        repaid = sum(t['amount'] for t in info['repayments'])
        received = sum(t['amount'] for t in info['setups'])
        loans_summary.append({
            'loan_id': lid,
            'owner': info['owner'],
            'received': received,
            'repaid': repaid,
            'outstanding_estimate': max(0, received - repaid) if received else None,  # only meaningful if we saw the loan setup
            'repayment_count': len(info['repayments']),
            'avg_payment': repaid / max(1, len(info['repayments'])),
            'last_payment_date': max((t['bill_date'] for t in info['repayments']), default=''),
        })
    loans_total_repaid = sum(l['repaid'] for l in loans_summary)
    loans_total_received = sum(l['received'] for l in loans_summary)

    # Category donut data (top 8 by total spending excluding transfers + earnings)
    cat_totals_dict = defaultdict(float)
    for t in cls:
        if t['status'] in ('transfer','earnings','rental_income','rent','landlord_rent'): continue
        cat_totals_dict[t['category'] or 'unknown'] += t['amount']
    top_cats = sorted(cat_totals_dict.items(), key=lambda x:-x[1])[:8]
    other_cats_total = sum(v for _,v in sorted(cat_totals_dict.items(), key=lambda x:-x[1])[8:])
    if other_cats_total > 0:
        top_cats.append(('Other', other_cats_total))

    # Monthly survive estimate (shared spend per month, 6 months)
    months_count = MONTHS_COUNT
    avg_shared_per_month = shared_total_with_rent / months_count
    avg_total_per_month = (totals['person_a']+totals['person_b']+assumed_rent_total) / months_count
    avg_personal = {o: sum(t['amount'] for t in personal_items[o])/months_count for o in ('person_a','person_b')}

    # Future projection baselines (monthly averages from current period)
    baseline = {
        'person_a_salary_pm':  earnings_person_a_salary / months_count,
        'person_b_salary_pm': earnings_person_b_salary / max(1, len([t for t in person_b_credits if t.get('earnings_kind')=='salary'])),  # use actual months observed
        'rental_income_pm': RENTAL_INCOME_AMOUNT,  # 2,100
        'person_a_other_credits_pm': earnings_person_a_other / months_count,
        'rent_pm': RENT_PER_MONTH_TOTAL,
        'utilities_pm': (sum(utilities[t]['person_a']+utilities[t]['person_b'] for t in utilities)) / months_count,
        'groceries_pm': sum(t['amount'] for t in shared_items if t.get('tag') in ('grocery','grocery_or_household')) / months_count,
        'restaurants_pm': sum(t['amount'] for t in restaurants) / months_count,
        'vacations_pm': sum(t['amount'] for t in vacations) / months_count,
        'cars_pm': (car_totals['person_a'] + car_totals['person_b']) / months_count,
        'business_pm': business_total / months_count,
        'subs_gym_pm': (subs_total + gyms_total) / months_count,
        'furniture_pm': furniture_total / months_count,
        'loans_pm': loans_total_repaid / months_count,
        'other_personal_pm': max(0, (totals['person_a']+totals['person_b']) / months_count - (
            sum(utilities[t]['person_a']+utilities[t]['person_b'] for t in utilities) / months_count
            + sum(t['amount'] for t in shared_items if t.get('tag') in ('grocery','grocery_or_household')) / months_count
            + sum(t['amount'] for t in restaurants) / months_count
            + sum(t['amount'] for t in vacations) / months_count
            + (car_totals['person_a'] + car_totals['person_b']) / months_count
            + business_total / months_count
            + (subs_total + gyms_total) / months_count
            + furniture_total / months_count
        )),
    }

    # Sort flagged by amount
    flagged_sorted = sorted(flagged, key=lambda x: -x['amount'])

    # ─── HTML ────────────────────────────────────────────────────────────
    report = render.build_html(
        totals=totals,
        by_status=by_status,
        by_tag=by_tag,
        by_month=by_month,
        utilities=utilities,
        flagged=flagged_sorted,
        shared_items=shared_items,
        personal_items=personal_items,
        transfers=transfers,
        rent_settlements=rent_settlements,
        rent_person_a_to_person_b=rent_person_a_to_person_b,
        rent_person_b_to_person_a=rent_person_b_to_person_a,
        assumed_rent_total=assumed_rent_total,
        assumed_rent_each=assumed_rent_each,
        landlord_payments=landlord_payments,
        person_a_landlord_months=person_a_landlord_months,
        person_a_landlord_count=person_a_landlord_count,
        person_b_landlord_count=person_b_landlord_count,
        person_a_landlord_paid=person_a_landlord_paid,
        person_b_landlord_paid_inferred=person_b_landlord_paid_inferred,
        rent_remaining_owed_to_person_a=rent_remaining_owed_to_person_a,
        rent_strict_owed_to_person_a=rent_strict_owed_to_person_a,
        person_a_rent_contribution=person_a_rent_contribution,
        shared_cards_person_a=shared_cards_person_a,
        shared_cards_person_b=shared_cards_person_b,
        shared_cards_total=shared_cards_total,
        shared_person_a_total=shared_person_a_total,
        shared_person_b_total=shared_person_b_total,
        shared_total=shared_total_with_rent,
        each_should_pay=each_should_pay,
        delta=delta_cards,
        final_delta=final_delta,
        bank_credits=bank_credits,
        bank_debits=bank_debits,
        earnings_person_a_total=earnings_person_a_total,
        earnings_person_a_salary=earnings_person_a_salary,
        earnings_person_a_rental=earnings_person_a_rental,
        earnings_person_a_other=earnings_person_a_other,
        earnings_person_b_total=earnings_person_b_total,
        earnings_person_b_salary=earnings_person_b_salary,
        earnings_person_b_other=earnings_person_b_other,
        person_b_credits=person_b_credits,
        person_b_landlord_months=person_b_landlord_months,
        rental_incomes=rental_incomes,
        restaurants=restaurants,
        vacations=vacations,
        rest_total=rest_total,
        rest_by_owner=rest_by_owner,
        vac_total=vac_total,
        vac_by_owner=vac_by_owner,
        cars=cars,
        car_totals=car_totals,
        car_breakdown_r=car_breakdown_r,
        car_breakdown_a=car_breakdown_a,
        subs=subs,
        gyms=gyms,
        subs_total=subs_total,
        gyms_total=gyms_total,
        furniture=furniture,
        furniture_total=furniture_total,
        furniture_by_owner=furniture_by_owner,
        business=business,
        business_total=business_total,
        business_by_owner=business_by_owner,
        loans_summary=loans_summary,
        loans_raw=loans_raw,
        loans_total_repaid=loans_total_repaid,
        loans_total_received=loans_total_received,
        baseline=baseline,
        top_cats=top_cats,
        avg_shared_per_month=avg_shared_per_month,
        avg_total_per_month=avg_total_per_month,
        avg_personal=avg_personal,
        months_count=months_count,
    )
    out = OUTDIR / C.REPORT_NAME
    out.write_text(report, encoding='utf-8')
    print(f'wrote {out}', file=sys.stderr)

if __name__ == '__main__':
    main()
